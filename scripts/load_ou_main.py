#!/usr/bin/env python3
"""
Загрузка листов «АРМ» и «Принтеры» из файла «Основной Учет.xlsx» в БД ias_vniic, схема tech_accounting.
Парсинг строго по docs/import_ou/РЕГЛАМЕНТ_ПАРСИНГА.md.

- Лист АРМ: одна строка → запись equipment (ПК/только монитор) + отдельные записи equipment на каждый
  монитор из столбца «Монитор» (МЦ.04-mon-XXX) и на каждый ИБП из «Другая техника» (МЦ.04-ups-XXX).
- Лист Принтеры: одна строка → одна запись equipment (Принтер/МФУ).
"""
import os
import re
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    import psycopg2
except ImportError as e:
    print("Требуются: pip install openpyxl psycopg2-binary")
    raise SystemExit(1) from e

PROJECT_ROOT = Path(__file__).resolve().parent.parent
EXCEL_PATH = os.environ.get("OU_MAIN_EXCEL", str(PROJECT_ROOT / "Основной Учет.xlsx"))
DB_HOST = os.environ.get("PGHOST", "localhost")
DB_PORT = os.environ.get("PGPORT", "5432")
DB_NAME = os.environ.get("PGDATABASE", "ias_vniic")
DB_USER = os.environ.get("PGUSER", "postgres")
DB_PASSWORD = os.environ.get("PGPASSWORD", "12345")
SCHEMA = "tech_accounting"


def norm(v):
    if v is None:
        return None
    if isinstance(v, str):
        return v.strip() or None
    return v


def first_line_only(s):
    """При \\n в Пользователе брать первую строку."""
    if not s:
        return None
    s = str(s).strip()
    return s.split("\n")[0].strip() if "\n" in s else s


def inv_first_line(s):
    """Для № системн. блока при \\n — первая подстрока."""
    if not s:
        return None
    s = str(s).strip()
    return s.split("\n")[0].strip() if "\n" in s else s


def disk_value(s):
    """Диск: разделители \\n и запятая (полный файл); в БД одна запись — объединяем в один value_text."""
    if not s:
        return None
    s = str(s).strip()
    parts = re.split(r"[\n,]+", s)
    parts = [p.strip() for p in parts if p.strip()]
    return "; ".join(parts) if parts else None


def monitor_parts(s):
    """Монитор: разбить по \\n и '; ' для списка моделей (для part_char_values и для отдельных equipment)."""
    if not s:
        return []
    s = str(s).strip()
    if ";" in s:
        parts = [p.strip() for p in s.split(";") if p.strip()]
    else:
        parts = [p.strip() for p in s.split("\n") if p.strip()]
    return parts


def parse_date(val):
    """Год (int) -> 01.01.YYYY; DD.MM.YYYY -> date; иначе None."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, (int, float)):
        y = int(val)
        if 1990 <= y <= 2030:
            return datetime(y, 1, 1).date()
        return None
    s = str(val).strip()
    if not s or "(" in s or "б/у" in s.lower():
        return None
    m = re.match(r"^(\d{1,2})\.(\d{1,2})\.(\d{4})$", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return datetime(y, mo, d).date()
        except ValueError:
            return None
    if re.match(r"^\d{4}$", s):
        return datetime(int(s), 1, 1).date()
    return None


def location_name(val):
    """Помещение: число -> строка."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return str(int(val))
    return str(val).strip() or None


def extract_ups_list(text):
    """Из «Другая техника» извлечь подстроки вида «ИБП <марка и модель>»; вернуть список name (без префикса ИБП )."""
    if not text:
        return []
    s = str(text).strip()
    result = []
    for part in re.split(r"[\n]+", s):
        part = part.strip()
        if part.startswith("ИБП "):
            result.append(part[4:].strip())
    return result


def load_sheet(path, sheet_name, max_rows=None):
    """Читает лист Excel; max_rows=None — все строки."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    headers = [str(c).strip() if c else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    max_row = (2 + max_rows) if max_rows else ws.max_row
    rows = list(ws.iter_rows(min_row=2, max_row=max_row, values_only=True))
    wb.close()
    return headers, rows


def get_col_index(headers, names):
    for name in names:
        for i, h in enumerate(headers):
            if h and name in h:
                return i
    return None


def ensure_spr_parts_chars(cur, parts_chars):
    """parts_chars = {"ЦП": ["Модель"], "ОЗУ": ["Объём"], ...}. Возвращает dict part_name -> id, char_name -> id."""
    part_ids = {}
    char_ids = {}
    for part_name, char_list in parts_chars.items():
        cur.execute("SELECT id FROM spr_parts WHERE name = %s", (part_name,))
        r = cur.fetchone()
        if not r:
            cur.execute("INSERT INTO spr_parts (name) VALUES (%s) RETURNING id", (part_name,))
            part_ids[part_name] = cur.fetchone()[0]
        else:
            part_ids[part_name] = r[0]
        for char_name in char_list:
            cur.execute("SELECT id FROM spr_chars WHERE name = %s", (char_name,))
            r = cur.fetchone()
            if not r:
                cur.execute("INSERT INTO spr_chars (name) VALUES (%s) RETURNING id", (char_name,))
                char_ids[char_name] = cur.fetchone()[0]
            else:
                char_ids[char_name] = r[0]
    return part_ids, char_ids


def load_arm_sheet(conn, cur, path, status_id, role_id, parts, chars, seen_inv, mon_counter, ups_counter, errors):
    headers, rows = load_sheet(path, "АРМ")
    idx_user = get_col_index(headers, ["Пользователь"])
    idx_room = get_col_index(headers, ["Помещение"])
    idx_cpu = get_col_index(headers, ["ЦП"])
    idx_ram = get_col_index(headers, ["ОЗУ"])
    idx_disk = get_col_index(headers, ["Диск"])
    idx_sb = get_col_index(headers, ["Системный блок"])
    idx_inv_sb = get_col_index(headers, ["№ системн. блока"])
    idx_date_sb = get_col_index(headers, ["Дата закупки системного блока"])
    idx_monitor = get_col_index(headers, ["Монитор"])
    idx_inv_mon = get_col_index(headers, ["№ монитора"])
    idx_date_mon = get_col_index(headers, ["Дата закупки монитора"])
    idx_hostname = get_col_index(headers, ["Имя"])
    idx_ip = get_col_index(headers, ["IP адрес"])
    idx_os = get_col_index(headers, ["ОС"])
    idx_av = get_col_index(headers, ["Антивирус"])
    idx_other = get_col_index(headers, ["Другая техника"])
    idx_note = get_col_index(headers, ["Примечание"])

    def get_or_create_location(name):
        name = location_name(name)
        if not name:
            return None
        cur.execute("SELECT id FROM locations WHERE name = %s", (name,))
        r = cur.fetchone()
        if r:
            return r[0]
        cur.execute(
            "INSERT INTO locations (name, location_type) VALUES (%s, 'кабинет') RETURNING id",
            (name,),
        )
        return cur.fetchone()[0]

    def get_or_create_user(full_name):
        full_name = first_line_only(full_name)
        if not full_name:
            return None
        cur.execute("SELECT id FROM users WHERE full_name = %s AND is_deleted = FALSE", (full_name,))
        r = cur.fetchone()
        if r:
            return r[0]
        cur.execute("INSERT INTO users (full_name) VALUES (%s) RETURNING id", (full_name,))
        uid = cur.fetchone()[0]
        if role_id:
            cur.execute(
                "INSERT INTO user_roles (user_id, role_id) VALUES (%s, %s) ON CONFLICT DO NOTHING",
                (uid, role_id),
            )
        return uid

    def add_pcv(equip_id, part_name, char_name, value):
        if value is None or not str(value).strip():
            return
        p_id = parts.get(part_name)
        c_id = chars.get(char_name)
        if p_id is None or c_id is None:
            return
        cur.execute(
            """INSERT INTO part_char_values (equipment_id, part_id, char_id, value_text)
               VALUES (%s, %s, %s, %s)
               ON CONFLICT (equipment_id, part_id, char_id) DO UPDATE SET value_text = EXCLUDED.value_text""",
            (equip_id, p_id, c_id, str(value).strip()[:5000]),
        )

    def insert_equipment(inv, name, eq_type, status_id, user_id, location_id, purchase_date, description):
        cur.execute(
            """INSERT INTO equipment (inventory_number, name, equipment_type, status_id, responsible_user_id, location_id, purchase_date, description)
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
               ON CONFLICT (inventory_number) DO UPDATE SET
                 name = EXCLUDED.name,
                 equipment_type = EXCLUDED.equipment_type,
                 status_id = EXCLUDED.status_id,
                 responsible_user_id = EXCLUDED.responsible_user_id,
                 location_id = EXCLUDED.location_id,
                 purchase_date = EXCLUDED.purchase_date,
                 description = EXCLUDED.description
               RETURNING id""",
            (
                inv,
                (name or inv)[:200],
                (eq_type or "Системный блок")[:100],
                status_id,
                user_id,
                location_id,
                purchase_date,
                description,
            ),
        )
        return cur.fetchone()[0]

    loaded = 0
    for row_num, row in enumerate(rows, start=2):
        def cell(i):
            return norm(row[i]) if i is not None and i < len(row) else None

        inv_sb = inv_first_line(cell(idx_inv_sb))
        inv_mon = cell(idx_inv_mon)
        has_sb = bool(inv_sb or cell(idx_sb))
        if has_sb and not inv_sb:
            inv_sb = inv_mon or f"МЦ.04-ROW-{row_num}"
        if not has_sb:
            inv_sb = inv_first_line(inv_mon) if inv_mon else f"МЦ.04-MON-{row_num}"
        if not inv_sb:
            errors.append(f"АРМ строка {row_num}: нет инв. номера, пропуск")
            continue
        if inv_sb in seen_inv:
            inv_sb = f"{inv_sb}-строка{row_num}"
        seen_inv.add(inv_sb)

        loc_name = location_name(cell(idx_room))
        if not loc_name:
            errors.append(f"АРМ строка {row_num}: нет помещения, пропуск")
            continue
        location_id = get_or_create_location(loc_name)
        if not location_id:
            continue
        user_id = get_or_create_user(row[idx_user] if idx_user is not None else None)

        name_equip = cell(idx_sb)
        equipment_type = "Системный блок"
        if name_equip:
            if "моноблок" in name_equip.lower():
                equipment_type = "Моноблок"
            elif "ноутбук" in name_equip.lower():
                equipment_type = "Ноутбук"
        if not name_equip and cell(idx_monitor):
            name_equip = "; ".join(monitor_parts(cell(idx_monitor))) or "Монитор"
            equipment_type = "Монитор"
        if not name_equip:
            name_equip = inv_sb
        if name_equip and "\n" in name_equip:
            name_equip = name_equip.replace("\n", "; ")

        purchase_date = parse_date(cell(idx_date_sb)) if idx_date_sb is not None else None

        desc_parts = []
        other_text = cell(idx_other)
        ups_list = extract_ups_list(other_text)
        if other_text:
            remaining = other_text
            for u in ups_list:
                remaining = remaining.replace("ИБП " + u, "").strip()
            remaining = re.sub(r"\n+", " ", remaining).strip()
            if remaining:
                desc_parts.append(remaining)
        if idx_note is not None and cell(idx_note):
            desc_parts.append(cell(idx_note))
        description = "; ".join(desc_parts) if desc_parts else None

        try:
            equip_id = insert_equipment(
                inv_sb, name_equip, equipment_type, status_id, user_id, location_id, purchase_date, description
            )
        except psycopg2.IntegrityError as e:
            errors.append(f"АРМ строка {row_num}: {e}")
            raise

        if idx_cpu is not None:
            add_pcv(equip_id, "ЦП", "Модель", cell(idx_cpu))
        if idx_ram is not None:
            add_pcv(equip_id, "ОЗУ", "Объём", cell(idx_ram))
        if idx_disk is not None:
            add_pcv(equip_id, "Накопитель", "Модель", disk_value(row[idx_disk]) if idx_disk < len(row) else None)
        if idx_monitor is not None and cell(idx_monitor):
            monitor_str = "; ".join(monitor_parts(cell(idx_monitor)))
            add_pcv(equip_id, "Монитор", "Модель", monitor_str)
        if idx_inv_mon is not None:
            add_pcv(equip_id, "Монитор", "№ монитора", cell(idx_inv_mon))
        if idx_hostname is not None:
            add_pcv(equip_id, "ПК", "Имя ПК", cell(idx_hostname))
        if idx_ip is not None:
            add_pcv(equip_id, "ПК", "IP адрес", cell(idx_ip))
        if idx_os is not None:
            add_pcv(equip_id, "ПК", "ОС", cell(idx_os))
        if idx_av is not None:
            add_pcv(equip_id, "ПК", "Антивирус", cell(idx_av))

        loaded += 1

        # Отдельные записи equipment на каждый монитор только для строки-ПК (есть СБ); «только монитор» — уже одна запись выше.
        monitor_models = monitor_parts(cell(idx_monitor)) if has_sb else []
        is_monoblock = name_equip and "моноблок" in name_equip.lower()
        for model in monitor_models:
            if not model or model.lower() == "моноблок":
                continue
            mon_counter[0] += 1
            inv_mon_eq = f"МЦ.04-mon-{mon_counter[0]}"
            if inv_mon_eq in seen_inv:
                inv_mon_eq = f"{inv_mon_eq}-строка{row_num}"
            seen_inv.add(inv_mon_eq)
            try:
                insert_equipment(inv_mon_eq, model, "Монитор", status_id, user_id, location_id, None, None)
            except psycopg2.IntegrityError as e:
                errors.append(f"АРМ строка {row_num} монитор {model}: {e}")
                raise

        for ups_name in ups_list:
            ups_counter[0] += 1
            inv_ups = f"МЦ.04-ups-{ups_counter[0]}"
            if inv_ups in seen_inv:
                inv_ups = f"{inv_ups}-строка{row_num}"
            seen_inv.add(inv_ups)
            try:
                insert_equipment(inv_ups, ups_name, "ИБП", status_id, user_id, location_id, None, None)
            except psycopg2.IntegrityError as e:
                errors.append(f"АРМ строка {row_num} ИБП {ups_name}: {e}")
                raise

    return loaded


def load_printers_sheet(conn, cur, path, status_id, role_id, parts, chars, seen_inv, errors):
    headers, rows = load_sheet(path, "Принтеры")
    idx_user = get_col_index(headers, ["Пользователь"])
    idx_room = get_col_index(headers, ["Помещение"])
    idx_org = get_col_index(headers, ["Оргтехника"])
    idx_inv = get_col_index(headers, ["Инвент. № принтера", "Инвент. №"])
    idx_date = get_col_index(headers, ["Дата закупки"])
    idx_ip = get_col_index(headers, ["IP адрес"])
    idx_note = get_col_index(headers, ["Примечания"])
    idx_note2 = get_col_index(headers, ["Примечание"])

    def get_or_create_location(name):
        name = location_name(name)
        if not name:
            return None
        cur.execute("SELECT id FROM locations WHERE name = %s", (name,))
        r = cur.fetchone()
        if r:
            return r[0]
        cur.execute(
            "INSERT INTO locations (name, location_type) VALUES (%s, 'кабинет') RETURNING id",
            (name,),
        )
        return cur.fetchone()[0]

    def get_or_create_user(full_name):
        full_name = first_line_only(full_name)
        if not full_name:
            return None
        cur.execute("SELECT id FROM users WHERE full_name = %s AND is_deleted = FALSE", (full_name,))
        r = cur.fetchone()
        if r:
            return r[0]
        cur.execute("INSERT INTO users (full_name) VALUES (%s) RETURNING id", (full_name,))
        uid = cur.fetchone()[0]
        if role_id:
            cur.execute(
                "INSERT INTO user_roles (user_id, role_id) VALUES (%s, %s) ON CONFLICT DO NOTHING",
                (uid, role_id),
            )
        return uid

    loaded = 0
    for row_num, row in enumerate(rows, start=2):
        def cell(i):
            return norm(row[i]) if i is not None and i < len(row) else None

        name_equip = first_line_only(cell(idx_org)) if idx_org is not None else None
        if not name_equip:
            errors.append(f"Принтеры строка {row_num}: нет оргтехники, пропуск")
            continue
        inv = first_line_only(cell(idx_inv)) if idx_inv is not None else None
        if not inv:
            inv = f"МЦ.04-принтер-{row_num}"
        if inv in seen_inv:
            inv = f"{inv}-строка{row_num}"
        seen_inv.add(inv)

        eq_type = "МФУ" if "мфу" in name_equip.lower() or "многофункциональн" in name_equip.lower() else "Принтер"
        loc_name = location_name(cell(idx_room))
        if not loc_name:
            errors.append(f"Принтеры строка {row_num}: нет помещения, пропуск")
            continue
        location_id = get_or_create_location(loc_name)
        if not location_id:
            continue
        user_id = get_or_create_user(row[idx_user] if idx_user is not None else None)
        purchase_date = parse_date(cell(idx_date)) if idx_date is not None else None
        desc_list = []
        if idx_note is not None and cell(idx_note):
            desc_list.append(first_line_only(cell(idx_note)))
        if idx_note2 is not None and idx_note2 != idx_note and cell(idx_note2):
            desc_list.append(cell(idx_note2))
        description = "; ".join(desc_list) if desc_list else None

        try:
            cur.execute(
                """INSERT INTO equipment (inventory_number, name, equipment_type, status_id, responsible_user_id, location_id, purchase_date, description)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                   ON CONFLICT (inventory_number) DO UPDATE SET
                     name = EXCLUDED.name,
                     equipment_type = EXCLUDED.equipment_type,
                     status_id = EXCLUDED.status_id,
                     responsible_user_id = EXCLUDED.responsible_user_id,
                     location_id = EXCLUDED.location_id,
                     purchase_date = EXCLUDED.purchase_date,
                     description = EXCLUDED.description
                   RETURNING id""",
                (inv, name_equip[:200], eq_type[:100], status_id, user_id, location_id, purchase_date, description),
            )
            equip_id = cur.fetchone()[0]
        except psycopg2.IntegrityError as e:
            errors.append(f"Принтеры строка {row_num}: {e}")
            raise

        if idx_ip is not None and cell(idx_ip):
            p_id = parts.get("Принтер")
            c_id = chars.get("IP адрес")
            if p_id is not None and c_id is not None:
                cur.execute(
                    """INSERT INTO part_char_values (equipment_id, part_id, char_id, value_text)
                       VALUES (%s, %s, %s, %s)
                       ON CONFLICT (equipment_id, part_id, char_id) DO UPDATE SET value_text = EXCLUDED.value_text""",
                    (equip_id, p_id, c_id, str(cell(idx_ip)).strip()[:5000]),
                )
        loaded += 1
    return loaded


def main():
    path = Path(EXCEL_PATH)
    if not path.exists():
        print(f"Файл не найден: {EXCEL_PATH}")
        print("Укажите путь переменной OU_MAIN_EXCEL или поместите «Основной Учет.xlsx» в корень проекта.")
        raise SystemExit(1)

    conn = psycopg2.connect(
        host=DB_HOST,
        port=DB_PORT,
        dbname=DB_NAME,
        user=DB_USER,
        password=DB_PASSWORD,
    )
    conn.autocommit = False
    cur = conn.cursor()
    cur.execute(f"SET search_path TO {SCHEMA}")

    cur.execute("SELECT id FROM dic_equipment_status WHERE status_code = 'in_use' LIMIT 1")
    row = cur.fetchone()
    if not row:
        print("В БД нет статуса in_use. Выполните scripts/create_ias_vniic.sql")
        conn.rollback()
        conn.close()
        raise SystemExit(1)
    status_id = row[0]

    cur.execute("SELECT id FROM roles WHERE role_code = 'user' LIMIT 1")
    r = cur.fetchone()
    role_id = r[0] if r else None

    parts_chars = {
        "ЦП": ["Модель"],
        "ОЗУ": ["Объём"],
        "Накопитель": ["Модель"],
        "Монитор": ["Модель", "№ монитора"],
        "ПК": ["Имя ПК", "IP адрес", "ОС", "Антивирус"],
        "Принтер": ["IP адрес"],
    }
    flat_parts = list(parts_chars.keys())
    flat_chars = []
    for cl in parts_chars.values():
        flat_chars.extend(cl)
    parts, chars = {}, {}
    for p in flat_parts:
        cur.execute("SELECT id FROM spr_parts WHERE name = %s", (p,))
        r = cur.fetchone()
        if not r:
            cur.execute("INSERT INTO spr_parts (name) VALUES (%s) RETURNING id", (p,))
            parts[p] = cur.fetchone()[0]
        else:
            parts[p] = r[0]
    for c in set(flat_chars):
        cur.execute("SELECT id FROM spr_chars WHERE name = %s", (c,))
        r = cur.fetchone()
        if not r:
            cur.execute("INSERT INTO spr_chars (name) VALUES (%s) RETURNING id", (c,))
            chars[c] = cur.fetchone()[0]
        else:
            chars[c] = r[0]

    errors = []
    seen_inv = set()
    mon_counter = [0]
    ups_counter = [0]

    try:
        loaded_arm = load_arm_sheet(
            conn, cur, str(path), status_id, role_id, parts, chars, seen_inv, mon_counter, ups_counter, errors
        )
        loaded_printers = load_printers_sheet(conn, cur, str(path), status_id, role_id, parts, chars, seen_inv, errors)
    except Exception:
        conn.rollback()
        cur.close()
        conn.close()
        raise

    conn.commit()
    cur.close()
    conn.close()

    print(f"Лист АРМ: загружено записей equipment (ПК/мониторы/ИБП): {loaded_arm}; дополнительно мониторов: {mon_counter[0]}, ИБП: {ups_counter[0]}")
    print(f"Лист Принтеры: загружено записей equipment: {loaded_printers}")
    if errors:
        for e in errors:
            print("  ", e)


if __name__ == "__main__":
    main()
