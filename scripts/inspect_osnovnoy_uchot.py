# -*- coding: utf-8 -*-
"""
Детальный разбор структуры файла «Основной Учет.xlsx».
Особое внимание — лист «АРМ»: заголовки, типы данных, неатомарные ячейки
(несколько значений в одной ячейке: перечисления носителей, комплектующих и т.д.).
"""

import re
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Установите openpyxl: pip install openpyxl")
    sys.exit(1)

PROJECT_ROOT = Path(__file__).resolve().parent.parent
EXCEL_PATH = PROJECT_ROOT / "Основной Учет.xlsx"

# Разделители, по которым в одной ячейке могут быть перечислены несколько значений
LIST_SEPARATORS = re.compile(r"[,;]|\s+и\s+|\s*[\n\r]+\s*")


def is_likely_multi(value):
    """Проверяет, похоже ли значение на перечисление (несколько элементов в одной ячейке)."""
    if value is None or not isinstance(value, str):
        return False
    s = value.strip()
    if not s or len(s) < 3:
        return False
    # Есть запятая, точка с запятой, " и ", перенос строки
    if "," in s or ";" in s or " и " in s or "\n" in s or "\r" in s:
        return True
    # Число перечислений через запятую (например "NVMe 512, HDD 1T")
    parts = LIST_SEPARATORS.split(s)
    return len([p for p in parts if p.strip()]) > 1


def describe_cell_value(v):
    """Краткое описание значения для отчёта."""
    if v is None:
        return "(пусто)"
    if isinstance(v, (int, float)):
        return f"[число] {v}"
    s = str(v).strip()
    if not s:
        return "(пусто)"
    length = len(s)
    preview = s[:80] + "…" if length > 80 else s
    multi = " [НЕАТОМАРНО: перечисление]" if is_likely_multi(s) else ""
    return f"{length} симв. | {repr(preview)}{multi}"


def inspect_sheet_arm(ws):
    """Детальный разбор листа АРМ: заголовки, строки, неатомарные ячейки."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"headers": [], "row_count": 0, "columns_detail": [], "sample_non_atomic": []}

    headers = [ (c or "").strip() for c in rows[0] ]
    # Нормализуем: столбец без имени — оставить индекс
    for i, h in enumerate(headers):
        if h == "":
            headers[i] = f"(столбец_{i+1})"

    row_count = len(rows) - 1  # без заголовка
    col_count = len(headers)

    # По каждому столбцу: есть ли непустые значения, есть ли неатомарные
    columns_detail = []
    for col_idx in range(col_count):
        col_name = headers[col_idx] if col_idx < len(headers) else f"Col{col_idx+1}"
        non_empty = 0
        non_atomic_count = 0
        examples_non_atomic = []
        for r in rows[1:]:
            if col_idx >= len(r):
                continue
            v = r[col_idx]
            if v is None or (isinstance(v, str) and not v.strip()):
                continue
            non_empty += 1
            if is_likely_multi(v):
                non_atomic_count += 1
                if len(examples_non_atomic) < 3:
                    examples_non_atomic.append(describe_cell_value(v))
        columns_detail.append({
            "index": col_idx + 1,
            "name": col_name,
            "non_empty": non_empty,
            "non_atomic_count": non_atomic_count,
            "examples_non_atomic": examples_non_atomic,
        })
        columns_detail[-1]["letter"] = get_column_letter(col_idx + 1)

    # Собрать примеры неатомарных ячеек по всему листу (столбец + значение)
    sample_non_atomic = []
    for row_idx, row in enumerate(rows[1:], start=2):
        for col_idx, v in enumerate(row):
            if col_idx >= len(headers):
                break
            if is_likely_multi(v):
                col_name = headers[col_idx] if col_idx < len(headers) else f"Col{col_idx+1}"
                sample_non_atomic.append({
                    "row": row_idx,
                    "col": col_name,
                    "letter": get_column_letter(col_idx + 1),
                    "preview": (str(v).strip()[:120] + "…") if len(str(v)) > 120 else str(v).strip(),
                })
                if len(sample_non_atomic) >= 15:
                    break
        if len(sample_non_atomic) >= 15:
            break

    return {
        "headers": headers,
        "row_count": row_count,
        "col_count": col_count,
        "columns_detail": columns_detail,
        "sample_non_atomic": sample_non_atomic,
        "first_data_rows": rows[1:6],  # первые 5 строк данных для контекста
    }


def main():
    if not EXCEL_PATH.exists():
        print(f"Файл не найден: {EXCEL_PATH}")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    print("=== Листы в файле ===")
    for i, name in enumerate(sheet_names):
        print(f"  [{i}] {name}")

    arm_sheet = None
    for name in sheet_names:
        if name.strip().upper() == "АРМ":
            arm_sheet = name
            break
    if not arm_sheet:
        # попробовать по частичному совпадению
        for name in sheet_names:
            if "арм" in name.lower():
                arm_sheet = name
                break
    if not arm_sheet:
        print("\nЛист «АРМ» не найден. Анализ первого листа по имени:", sheet_names[0] if sheet_names else "—")
        arm_sheet = sheet_names[0] if sheet_names else None

    if not arm_sheet:
        wb.close()
        sys.exit(1)

    ws = wb[arm_sheet]
    info = inspect_sheet_arm(ws)
    wb.close()

    print(f"\n=== Лист «{arm_sheet}» ===")
    print(f"Строк данных (без заголовка): {info['row_count']}")
    print(f"Столбцов: {info['col_count']}")
    print("\n--- Заголовки ---")
    for i, h in enumerate(info["headers"]):
        print(f"  {i+1}. {repr(h)}")

    print("\n--- Столбцы: заполненность и неатомарные значения ---")
    for c in info["columns_detail"]:
        na = c["non_atomic_count"]
        mark = " [ЕСТЬ НЕАТОМАРНЫЕ]" if na else ""
        print(f"  {c['letter']:>3} | {c['name'][:40]:<40} | непустых: {c['non_empty']}, неатомарных: {na}{mark}")
        for ex in c["examples_non_atomic"]:
            print(f"       Пример: {ex}")

    print("\n--- Примеры неатомарных ячеек (строка, столбец, содержимое) ---")
    for s in info["sample_non_atomic"]:
        print(f"  Строка {s['row']}, {s['col']} ({s['letter']}): {s['preview']}")

    print("\n--- Первые 5 строк данных (сырые значения по столбцам) ---")
    headers = info["headers"]
    for r in info["first_data_rows"]:
        row_preview = []
        for i, v in enumerate(r):
            if i >= len(headers):
                break
            desc = describe_cell_value(v)
            if len(desc) > 50:
                desc = desc[:50] + "…"
            row_preview.append(f"{headers[i][:12]}: {desc}")
        print("  | ".join(row_preview))
        print()

    # Вывод в файл для использования в документации
    out_path = PROJECT_ROOT / "docs" / "СТРУКТУРА_ОСНОВНОЙ_УЧЕТ_АРМ.txt"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("Структура листа «АРМ» файла «Основной Учет.xlsx» (автогенерация inspect_osnovnoy_uchot.py)\n\n")
        f.write("Заголовки: " + " | ".join(info["headers"]) + "\n\n")
        f.write("Столбцы с неатомарными значениями:\n")
        for c in info["columns_detail"]:
            if c["non_atomic_count"] > 0:
                f.write(f"  {c['name']}: неатомарных {c['non_atomic_count']}, примеры: {c['examples_non_atomic']}\n")
    print(f"\nКраткий отчёт записан в {out_path}")


if __name__ == "__main__":
    main()
